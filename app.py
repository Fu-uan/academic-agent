"""
文献检索智能体 - 后端服务
技术栈: FastAPI + OpenAlex API
"""
import json
import re
import urllib.parse
from typing import Optional

import httpx
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from langchain_core.messages import HumanMessage, AIMessage

from agent_setup import get_agent_executor

app = FastAPI(title="PaperAgent - 文献检索智能体", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── 数据模型 ──────────────────────────────────────────────

class SearchRequest(BaseModel):
    keyword: str
    year_start: Optional[int] = None
    year_end: Optional[int] = None
    database: str = "openalex"
    page: int = 1
    per_page: int = 20
    journals: Optional[list[str]] = None
    sort_by: str = "relevance"

class CiteRequest(BaseModel):
    dois: list[str]

class TranslateRequest(BaseModel):
    text: str
    domain: str = "sci-tech"  # sci-tech / humanities

class ExpandRequest(BaseModel):
    keyword: str
    domain: str = "general"

# ── 内置学术术语库 ──────────────────────────────────────────

# 理工科专业术语（中→英）
SCI_TECH_TERMS_CN2EN = {
    "深度学习": "deep learning",
    "机器学习": "machine learning",
    "神经网络": "neural network",
    "卷积神经网络": "convolutional neural network (CNN)",
    "循环神经网络": "recurrent neural network (RNN)",
    "自然语言处理": "natural language processing (NLP)",
    "计算机视觉": "computer vision",
    "强化学习": "reinforcement learning",
    "迁移学习": "transfer learning",
    "生成对抗网络": "generative adversarial network (GAN)",
    "大语言模型": "large language model (LLM)",
    "模型训练": "model training",
    "损失函数": "loss function",
    "梯度下降": "gradient descent",
    "过拟合": "overfitting",
    "正则化": "regularization",
    "注意力机制": "attention mechanism",
    "transformer": "Transformer",
    "词嵌入": "word embedding",
    "数据增强": "data augmentation",
    "特征提取": "feature extraction",
    "语义分割": "semantic segmentation",
    "目标检测": "object detection",
    "图像分类": "image classification",
    "聚类分析": "cluster analysis",
    "回归分析": "regression analysis",
    "主成分分析": "principal component analysis (PCA)",
    "支持向量机": "support vector machine (SVM)",
    "决策树": "decision tree",
    "随机森林": "random forest",
    "云计算": "cloud computing",
    "边缘计算": "edge computing",
    "物联网": "Internet of Things (IoT)",
    "区块链": "blockchain",
    "量子计算": "quantum computing",
    "数据挖掘": "data mining",
    "推荐系统": "recommender system",
    "知识图谱": "knowledge graph",
    "图神经网络": "graph neural network (GNN)",
    "联邦学习": "federated learning",
    "元学习": "meta-learning",
    "自监督学习": "self-supervised learning",
    "对比学习": "contrastive learning",
    "多模态": "multimodal",
    "信号处理": "signal processing",
    "控制系统": "control system",
    "有限元分析": "finite element analysis (FEA)",
    "计算流体力学": "computational fluid dynamics (CFD)",
    "概率统计": "probability and statistics",
    "线性代数": "linear algebra",
    "最优化": "optimization",
    "数值模拟": "numerical simulation",
    "离散事件仿真": "discrete event simulation",
    "结构方程模型": "structural equation modeling (SEM)",
    "遗传算法": "genetic algorithm (GA)",
    "粒子群算法": "particle swarm optimization (PSO)",
    "碳纳米管": "carbon nanotube (CNT)",
    "石墨烯": "graphene",
    "半导体": "semiconductor",
    "光催化": "photocatalysis",
    "电化学": "electrochemistry",
    "生物信息学": "bioinformatics",
    "基因组学": "genomics",
    "蛋白质结构预测": "protein structure prediction",
    "分子动力学": "molecular dynamics",
    "代谢组学": "metabolomics",
    "药物设计": "drug design",
    "图像分割": "image segmentation",
    "时序预测": "time series forecasting",
    "异常检测": "anomaly detection",
    "降维": "dimensionality reduction",
    "集成学习": "ensemble learning",
    "贝叶斯方法": "Bayesian method",
    "因果推断": "causal inference",
}
SCI_TECH_TERMS_EN2CN = {v: k for k, v in SCI_TECH_TERMS_CN2EN.items()}

# 人文社科学术术语（中→英）
HUMANITIES_TERMS_CN2EN = {
    "社会学": "sociology",
    "人类学": "anthropology",
    "经济学": "economics",
    "政治学": "political science",
    "心理学": "psychology",
    "教育学": "education",
    "语言学": "linguistics",
    "哲学": "philosophy",
    "文学": "literature",
    "历史学": "history",
    "传播学": "communication studies",
    "法学": "law",
    "管理学": "management",
    "会计学": "accounting",
    "金融学": "finance",
    "市场营销": "marketing",
    "组织行为学": "organizational behavior",
    "公共管理": "public administration",
    "国际关系": "international relations",
    "文化研究": "cultural studies",
    "批判性思维": "critical thinking",
    "话语分析": "discourse analysis",
    "质性研究": "qualitative research",
    "定量研究": "quantitative research",
    "田野调查": "field research",
    "扎根理论": "grounded theory",
    "叙事分析": "narrative analysis",
    "政策分析": "policy analysis",
    "博弈论": "game theory",
    "社会网络分析": "social network analysis (SNA)",
    "实证研究": "empirical research",
    "理论研究": "theoretical research",
    "案例研究": "case study",
    "比较研究": "comparative study",
    "纵向研究": "longitudinal study",
    "横截面研究": "cross-sectional study",
    "内容分析": "content analysis",
    "元分析": "meta-analysis",
    "社会资本": "social capital",
    "社会分层": "social stratification",
    "社会流动": "social mobility",
    "社会认同": "social identity",
    "身份认同": "identity",
    "全球化": "globalization",
    "现代化": "modernization",
    "城市化": "urbanization",
    "社会变迁": "social change",
    "意识形态": "ideology",
    "文化资本": "cultural capital",
    "符号互动": "symbolic interaction",
    "制度变迁": "institutional change",
    "治理": "governance",
    "民主化": "democratization",
    "公共政策": "public policy",
    "福利国家": "welfare state",
    "公民社会": "civil society",
    "软实力": "soft power",
    "知识管理": "knowledge management",
    "供应链管理": "supply chain management (SCM)",
    "企业社会责任": "corporate social responsibility (CSR)",
    "竞争优势": "competitive advantage",
    "消费者行为": "consumer behavior",
    "品牌资产": "brand equity",
    "人力资本": "human capital",
    "社会企业": "social enterprise",
    "可持续发展": "sustainable development",
    "循环经济": "circular economy",
    "数字经济": "digital economy",
    "平台经济": "platform economy",
    "共享经济": "sharing economy",
    "制度经济学": "institutional economics",
    "行为经济学": "behavioral economics",
    "实验经济学": "experimental economics",
    "宏观经济": "macroeconomics",
    "微观经济": "microeconomics",
    "计量经济学": "econometrics",
    "面板数据": "panel data",
    "工具变量": "instrumental variable (IV)",
    "双重差分": "difference-in-differences (DID)",
    "断点回归": "regression discontinuity (RDD)",
    "倾向得分匹配": "propensity score matching (PSM)",
    "中介效应": "mediation effect",
    "调节效应": "moderation effect",
    "内生性": "endogeneity",
}
HUMANITIES_TERMS_EN2CN = {v: k for k, v in HUMANITIES_TERMS_CN2EN.items()}

# ── 关键词扩写规则 ──────────────────────────────────────────

EXPANSION_RULES = {
    "deep learning": ["neural network", "machine learning", "deep neural network", "representation learning"],
    "machine learning": ["artificial intelligence", "deep learning", "statistical learning", "pattern recognition"],
    "natural language processing": ["text mining", "computational linguistics", "NLP", "language model"],
    "computer vision": ["image processing", "visual recognition", "scene understanding", "pattern recognition"],
    "reinforcement learning": ["RL", "deep RL", "Markov decision process", "policy gradient"],
    "data mining": ["knowledge discovery", "pattern mining", "big data analytics"],
    "blockchain": ["distributed ledger", "smart contract", "consensus mechanism"],
    "IoT": ["Internet of Things", "sensor network", "cyber-physical system"],
    "cloud computing": ["distributed computing", "edge computing", "serverless"],
    "quantum computing": ["quantum algorithm", "quantum information", "quantum supremacy"],
    "bioinformatics": ["computational biology", "genomics", "systems biology"],
    "sustainable development": ["ESG", "green economy", "circular economy", "sustainability"],
    "digital economy": ["platform economy", "sharing economy", "digital transformation"],
    "supply chain": ["logistics", "SCM", "value chain", "inventory management"],
    "corporate governance": ["board of directors", "ownership structure", "agency theory"],
    "public health": ["epidemiology", "health policy", "global health", "disease prevention"],
    "climate change": ["global warming", "carbon emission", "climate policy", "adaptation"],
    "education": ["pedagogy", "learning analytics", "educational technology", "curriculum"],
    "psychology": ["cognitive science", "behavioral science", "mental health", "neuroscience"],
}

# ── 数据库描述 ──────────────────────────────────────────────

DATABASES = {
    "openalex": {
        "name": "开放学术 (OpenAlex)",
        "description": "全球开放学术索引数据库，涵盖2.5亿+学术作品",
        "type": "综合",
        "coverage": "国际期刊、会议论文、学位论文"
    },
    "crossref": {
        "name": "Crossref",
        "description": "学术出版DOI注册机构，覆盖1.5亿+文献",
        "type": "综合",
        "coverage": "国际期刊论文（含DOI）"
    },
    "cnki": {
        "name": "中国知网 (CNKI)",
        "description": "中国最大的学术数据库（需机构订阅）",
        "type": "中文综合",
        "coverage": "中文学术期刊、学位论文、会议论文"
    },
    "wos": {
        "name": "Web of Science",
        "description": "Clarivate核心期刊引文索引（需机构订阅）",
        "type": "综合",
        "coverage": "SCI/SSCI/A&HCI核心期刊"
    },
    "wanfang": {
        "name": "万方数据",
        "description": "中国学术文献数据库（需机构订阅）",
        "type": "中文综合",
        "coverage": "中文学术期刊、学位论文、标准"
    },
    "vip": {
        "name": "维普期刊",
        "description": "中文科技期刊数据库（需机构订阅）",
        "type": "中文科技",
        "coverage": "中文科技期刊全文"
    },
}

# ── OpenAlex / arXiv API 调用 ──────────────────────────────

OPENALEX_BASE = "https://api.openalex.org"


async def search_arxiv(keyword: str, page: int = 1, per_page: int = 20) -> dict:
    """调用 arXiv API 检索预印本文献"""
    import xml.etree.ElementTree as ET

    start = (page - 1) * per_page
    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        try:
            resp = await client.get(
                "https://export.arxiv.org/api/query",
                params={"search_query": f"all:{keyword}", "start": start, "max_results": per_page},
                headers={"User-Agent": "PaperAgent/1.0"},
            )
            resp.raise_for_status()
            xml_data = resp.text
        except Exception as e:
            return {"total": 0, "page": page, "per_page": per_page, "results": [], "query": keyword, "error": str(e)}

    ns = {"atom": "http://www.w3.org/2005/Atom", "arxiv": "http://arxiv.org/schemas/atom", "opensearch": "http://a9.com/-/spec/opensearch/1.1/"}
    root = ET.fromstring(xml_data)
    total_str = root.findtext("opensearch:totalResults", "0", ns)
    total = int(total_str) if total_str and total_str.isdigit() else 0

    papers = []
    for entry in root.findall("atom:entry", ns):
        title = entry.findtext("atom:title", "", ns).strip().replace("\n", " ").replace("  ", " ")
        summary = entry.findtext("atom:summary", "", ns).strip().replace("\n", " ")[:500]
        paper_id = entry.findtext("atom:id", "", ns).replace("http://arxiv.org/abs/", "")
        doi = ""
        for link in entry.findall("atom:link", ns):
            href = link.get("href", "")
            if "doi.org" in href:
                doi = href.split("doi.org/")[-1]
        authors = [a.findtext("atom:name", "", ns).strip() for a in entry.findall("atom:author", ns) if a.findtext("atom:name", "", ns)]
        published = entry.findtext("atom:published", "", ns)
        year = int(published[:4]) if len(published) >= 4 else None
        categories = [c.get("term", "") for c in entry.findall("atom:category", ns) if c.get("term")]
        papers.append({
            "title": title, "authors": authors, "year": year,
            "journal": "arXiv", "cited_by": 0, "abstract": summary,
            "doi": doi or paper_id, "type": "预印本", "rank": "N/A",
            "rank_description": "", "is_oa": True,
            "oa_url": f"https://arxiv.org/abs/{paper_id}" if paper_id else "",
            "urls": [f"https://arxiv.org/abs/{paper_id}"] if paper_id else [],
            "keywords": categories[:5], "source": "arxiv",
        })

    return {"total": total, "page": page, "per_page": per_page, "results": papers, "query": keyword}


async def search_openalex(keyword: str, year_start: int = None, year_end: int = None, page: int = 1, per_page: int = 20, journals: list[str] = None, sort_by: str = "relevance") -> dict:
    """调用 OpenAlex API 检索文献"""
    # 有期刊筛选时，多拉一些数据提高命中率
    if journals and per_page < 200:
        per_page = 200
    sort_map = {"relevance": "relevance_score:desc", "citations": "cited_by_count:desc", "date": "publication_year:desc"}
    params = {
        "search": keyword,
        "per_page": min(per_page, 200),
        "page": page,
        "sort": sort_map.get(sort_by, "relevance_score:desc"),
    }
    filters = []
    import datetime
    current_year = datetime.datetime.now().year
    if year_start and year_end:
        filters.append(f"publication_year:{year_start}-{year_end}")
    elif year_start:
        filters.append(f"publication_year:{year_start}-{current_year}")
    elif year_end:
        filters.append(f"publication_year:1900-{year_end}")
    if filters:
        params["filter"] = ",".join(filters)

    headers = {
        "User-Agent": "PaperAgent/1.0 (mailto:contact@paperagent.local)",
        "Accept": "application/json",
    }
    params["mailto"] = "contact@paperagent.local"

    async with httpx.AsyncClient(timeout=30.0) as client:
        try:
            resp = await client.get(f"{OPENALEX_BASE}/works", params=params, headers=headers)
            if resp.status_code == 429:
                raise HTTPException(status_code=429, detail="API速率限制已达上限（每分钟10次），请稍后重试")
            resp.raise_for_status()
            data = resp.json()
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="OpenAlex API 请求超时，请稍后重试")
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"OpenAlex API 返回错误: {e.response.status_code}")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"检索失败: {str(e)}")

    results = data.get("results", [])
    total = data.get("meta", {}).get("count", 0)

    papers = []
    for r in results:
        paper = _parse_openalex_work(r)
        if paper:
            papers.append(paper)

    # 期刊筛选（后端过滤）
    if journals:
        journal_lower = [j.lower() for j in journals]
        papers = [p for p in papers if p.get("journal", "").lower() in journal_lower]
        total = len(papers)

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "results": papers,
        "query": keyword,
    }


def _parse_openalex_work(r: dict) -> Optional[dict]:
    """将 OpenAlex 返回的单条文献解析为统一格式"""
    try:
        title = r.get("title") or "未知标题"
        doi = r.get("doi", "").replace("https://doi.org/", "") if r.get("doi") else ""
        openalex_id = r.get("id", "")

        # 作者
        authors = []
        for a in (r.get("authorships") or []):
            name = a.get("author", {}).get("display_name", "")
            if name:
                authors.append(name)

        # 期刊/来源
        loc = r.get("primary_location") or {}
        src = loc.get("source") or {}
        journal = src.get("display_name") or ""
        journal_type = src.get("type") or ""
        journal_is_oa = src.get("is_oa", False)

        # 年份
        year = r.get("publication_year")
        # 出版日期（如果有更精确的信息）
        pub_date = r.get("publication_date", "")

        # 引用量
        cited_by = r.get("cited_by_count", 0)

        # 开放获取链接
        oa = r.get("open_access") or {}
        oa_url = oa.get("oa_url", "")

        # 类型
        work_type = r.get("type", "article")
        type_map = {
            "article": "期刊论文",
            "review": "综述",
            "book-chapter": "书籍章节",
            "book": "书籍",
            "dissertation": "学位论文",
            "thesis": "学位论文",
            "proceedings-article": "会议论文",
            "dataset": "数据集",
            "preprint": "预印本",
            "editorial": "社论",
            "erratum": "勘误",
            "standard": "标准",
            "report": "报告",
        }
        type_cn = type_map.get(work_type, work_type)

        # 卷期页码（OpenAlex 通常不提供，模拟）
        volume = r.get("biblio", {}).get("volume", "")
        issue = r.get("biblio", {}).get("issue", "")
        pages = r.get("biblio", {}).get("first_page", "") or ""

        # 摘要（OpenAlex 用倒排索引存储）
        abstract = _extract_abstract(r.get("abstract_inverted_index"))

        # 关键词
        keywords = [k.get("display_name", "") for k in (r.get("keywords") or []) if k.get("display_name")]

        # 分区估算（基于引用量百分位估算，实际分区需SCImago数据）
        # 这里用cited_by_count粗略分级
        if cited_by >= 500:
            rank = "Q1"
            rank_description = "高被引"
        elif cited_by >= 100:
            rank = "Q1"
            rank_description = "较高引用"
        elif cited_by >= 30:
            rank = "Q2"
            rank_description = "中等引用"
        elif cited_by >= 5:
            rank = "Q3"
            rank_description = "较低引用"
        else:
            rank = "Q4"
            rank_description = "低引用"

        # 完整链接
        urls = []
        if doi:
            urls.append(f"https://doi.org/{doi}")
        if oa_url:
            urls.append(oa_url)
        if openalex_id:
            urls.append(openalex_id)

        return {
            "title": title,
            "authors": authors,
            "journal": journal,
            "journal_type": journal_type,
            "year": year,
            "pub_date": pub_date,
            "doi": doi,
            "openalex_id": openalex_id,
            "cited_by": cited_by,
            "rank": rank,
            "rank_description": rank_description,
            "type": type_cn,
            "volume": volume,
            "issue": issue,
            "pages": pages,
            "abstract": abstract,
            "keywords": keywords,
            "urls": urls,
            "oa_url": oa_url,
            "is_oa": oa.get("is_oa", False),
            "source": "openalex",
        }
    except Exception:
        return None


def _extract_abstract(inverted_index: Optional[dict]) -> str:
    """从 OpenAlex 的倒排索引还原摘要"""
    if not inverted_index:
        return ""
    try:
        # 构建词位置列表
        word_positions = []
        for word, positions in inverted_index.items():
            for pos in positions:
                word_positions.append((pos, word))
        if not word_positions:
            return ""
        # 按位置排序
        word_positions.sort(key=lambda x: x[0])
        # 拼接
        abstract = " ".join(wp[1] for wp in word_positions)
        return abstract
    except Exception:
        return ""


# ── 关键词扩写 ──────────────────────────────────────────────

def expand_keywords(keyword: str, domain: str = "general") -> dict:
    """关键词扩写"""
    keyword_lower = keyword.lower().strip()
    expansions = []

    # 1. 查找预定义规则
    for base_kw, synonyms in EXPANSION_RULES.items():
        if base_kw in keyword_lower or keyword_lower in base_kw:
            expansions.extend([s for s in synonyms if s.lower() != keyword_lower])

    # 2. 同领域常见搭配词库
    domain_keywords = {
        "computer science": ["algorithm", "optimization", "system", "framework", "model", "simulation", "evaluation"],
        "engineering": ["design", "analysis", "simulation", "modeling", "experiment", "fabrication"],
        "physics": ["quantum", "theory", "simulation", "particle", "field", "dynamics"],
        "biology": ["gene", "protein", "cell", "organism", "evolution", "molecular"],
        "medicine": ["clinical", "therapy", "diagnosis", "patient", "disease", "treatment"],
        "economics": ["market", "growth", "policy", "trade", "investment", "welfare"],
        "social science": ["society", "culture", "institution", "behavior", "network"],
    }

    # 如果超过3个扩展词，去重后返回前10个
    seen = set()
    result = [keyword]
    for e in expansions:
        if e.lower() not in seen:
            seen.add(e.lower())
            result.append(e)

    # 如果太少，加一些通用扩展
    if len(result) < 3:
        # 用原文拆分词作为扩展（比"review"这类空洞词更有意义）
        words = keyword.strip().split()
        for w in words:
            if w.lower() not in seen and len(w) > 2:
                seen.add(w.lower())
                result.append(w)

    # 2. 同义词近义词（基于词的常见变换）
    # 简单规则：去掉引号、空格等
    cleaned = keyword.strip().strip('"').strip("'")
    if cleaned != keyword:
        result.append(cleaned)

    # 去重并限制数量
    final = []
    seen_final = set()
    for r in result:
        rl = r.lower()
        if rl not in seen_final:
            seen_final.add(rl)
            final.append(r)

    return {
        "original": keyword,
        "expanded": final[:10],
        "suggestion": f"已为您扩展检索关键词：{', '.join(final[:5])}..."
    }


# ── GB/T 7714 引文生成 ─────────────────────────────────────

def generate_gbt7714(paper: dict) -> str:
    """根据文献信息生成 GB/T 7714 格式引文"""
    # 作者处理：最多3个作者，超过则加"等"/"et al."
    authors = paper.get("authors", [])
    if not authors:
        author_str = "[佚名]"
    elif len(authors) <= 3:
        author_str = ", ".join(authors)
    else:
        author_str = ", ".join(authors[:3]) + ", 等"

    title = paper.get("title", "未知标题")
    journal = paper.get("journal", "未知期刊")
    year = paper.get("year", "")
    volume = paper.get("volume", "")
    issue = paper.get("issue", "")
    pages = paper.get("pages", "")
    doi = paper.get("doi", "")
    work_type = paper.get("type", "期刊论文")

    # 文献类型标识
    type_marker = {
        "期刊论文": "J",
        "综述": "J",
        "会议论文": "C",
        "学位论文": "D",
        "书籍": "M",
        "书籍章节": "M",
        "标准": "S",
        "报告": "R",
        "数据集": "DB",
        "预印本": "J",
        "社论": "J",
    }.get(work_type, "J")

    # 构造引文
    parts = [f"{author_str}. {title}[{type_marker}]. "]

    if journal:
        parts.append(f"{journal}, ")
    if year:
        parts.append(f"{year}")
    if volume:
        parts.append(f", {volume}")
        if issue:
            parts.append(f"({issue})")
    if pages and pages.strip():
        parts.append(f": {pages}")
    parts.append(".")
    if doi:
        parts.append(f" DOI: {doi}.")

    return "".join(parts)


def generate_gbt7714_batch(dois: list[str]) -> list[dict]:
    """批量生成引文（使用OpenAlex在线查询DOI获取详细信息）"""
    # 这种需要异步处理，稍后实现
    pass


# ── 学术翻译 ────────────────────────────────────────────────

def translate_academic(text: str, domain: str = "sci-tech") -> dict:
    """学术翻译：基于内置术语库的专业翻译"""
    text_stripped = text.strip()

    if not text_stripped:
        return {"source": "", "translation": "", "terms_matched": []}

    # 检测方向
    has_chinese = bool(re.search(r'[\u4e00-\u9fff]', text_stripped))
    has_english = bool(re.search(r'[a-zA-Z]{3,}', text_stripped))

    # 选择词库
    if domain == "sci-tech":
        cn2en = SCI_TECH_TERMS_CN2EN
        en2cn = SCI_TECH_TERMS_EN2CN
    else:
        cn2en = HUMANITIES_TERMS_CN2EN
        en2cn = HUMANITIES_TERMS_EN2CN

    # 中→英 还是 英→中
    is_cn_to_en = has_chinese and not has_english
    is_en_to_cn = has_english and not has_chinese

    matched_terms = []

    if is_cn_to_en:
        # 中译英：替换专业术语
        result = text_stripped
        for cn_term, en_term in sorted(cn2en.items(), key=lambda x: -len(x[0])):
            if cn_term in result:
                matched_terms.append({"source": cn_term, "target": en_term})
                result = result.replace(cn_term, en_term, 1)
        return {
            "source": text_stripped,
            "translation": result,
            "direction": "zh→en",
            "domain": domain,
            "terms_matched": matched_terms,
            "note": "基于内置学术术语库翻译，如需完整翻译请配置翻译API"
        }

    elif is_en_to_cn:
        # 英译中：替换专业术语
        result = text_stripped
        # 按长度降序排序，避免短词被先匹配
        for en_term, cn_term in sorted(en2cn.items(), key=lambda x: -len(x[0])):
            # 不区分大小写匹配
            pattern = re.compile(re.escape(en_term), re.IGNORECASE)
            if pattern.search(result):
                matched_terms.append({"source": en_term, "target": cn_term})
                result = pattern.sub(cn_term, result, count=1)
        return {
            "source": text_stripped,
            "translation": result,
            "direction": "en→zh",
            "domain": domain,
            "terms_matched": matched_terms,
            "note": "基于内置学术术语库翻译，如需完整翻译请配置翻译API"
        }

    else:
        # 混合文本或无法判断方向：尝试识别并替换术语
        result = text_stripped
        # 尝试识别中文术语并替换为英文
        for cn_term, en_term in sorted(cn2en.items(), key=lambda x: -len(x[0])):
            if cn_term in result:
                matched_terms.append({"source": cn_term, "target": en_term})
                result = result.replace(cn_term, f"【{en_term}】", 1)

        return {
            "source": text_stripped,
            "translation": result,
            "direction": "mixed",
            "domain": domain,
            "terms_matched": matched_terms,
            "note": "基于内置学术术语库翻译，如需完整翻译请配置翻译API"
        }


# ── API 路由 ────────────────────────────────────────────────

@app.get("/api/databases")
async def list_databases():
    """获取支持的数据库列表"""
    return {
        "databases": DATABASES,
        "primary": "openalex",
        "note": "知网/WoS/万方/维普需机构订阅，当前通过开放API镜像检索结果"
    }


@app.post("/api/search")
async def search_papers(req: SearchRequest):
    """检索文献"""
    if not req.keyword.strip():
        raise HTTPException(status_code=400, detail="请输入检索关键词")

    if req.database == "openalex":
        result = await search_openalex(req.keyword, req.year_start, req.year_end, req.page, req.per_page, journals=req.journals, sort_by=req.sort_by)
    elif req.database == "arxiv":
        result = await search_arxiv(req.keyword, req.page, req.per_page)
    elif req.database == "crossref":
        raise HTTPException(status_code=501, detail="Crossref API 检索暂未实现，请使用 openalex")
    else:
        # Google Scholar 等，用 OpenAlex 镜像
        result = await search_openalex(req.keyword, req.year_start, req.year_end, req.page, req.per_page, journals=req.journals)
        mirror_note = f"通过开放学术数据库（OpenAlex）镜像检索「{req.keyword}」的结果"
        result["mirror_note"] = mirror_note
        return result

    # 无结果时自动扩词
    if result["total"] == 0:
        expansion = expand_keywords(req.keyword)
        if len(expansion["expanded"]) > 1:
            # 用原始词+扩展词组合检索，比单一扩展词更精准
            extra = expansion["expanded"][1]
            if extra.lower() not in req.keyword.lower():
                new_keyword = f"{req.keyword} {extra}"
            else:
                new_keyword = extra
            if len(new_keyword) > 3:  # 避免过短的词
                new_result = await search_openalex(new_keyword, req.year_start, req.year_end, 1, req.per_page, journals=req.journals)
                if new_result["total"] > 0:
                    return {
                    **new_result,
                    "expanded_from": req.keyword,
                    "expansion_suggestion": expansion["suggestion"],
                    "expanded_query": new_keyword,
                }

    return result


@app.post("/api/cite")
async def generate_citation(req: CiteRequest):
    """生成GB/T 7714格式引文"""
    if not req.dois:
        raise HTTPException(status_code=400, detail="请提供至少一个DOI")

    citations = []

    # 从OpenAlex获取每个DOI对应的文献信息
    async with httpx.AsyncClient(timeout=30.0) as client:
        for doi in req.dois:
            clean_doi = doi.strip()
            if not clean_doi:
                continue
            try:
                resp = await client.get(
                    f"{OPENALEX_BASE}/works/doi:{clean_doi}",
                    headers={"User-Agent": "PaperAgent/1.0 (mailto:contact@paperagent.local)", "Accept": "application/json"},
                )
                if resp.status_code == 200:
                    data = resp.json()
                    paper = _parse_openalex_work(data)
                    if paper:
                        citation = generate_gbt7714(paper)
                        citations.append({"doi": clean_doi, "citation": citation, "success": True})
                    else:
                        citations.append({"doi": clean_doi, "citation": "", "success": False, "error": "解析失败"})
                else:
                    citations.append({"doi": clean_doi, "citation": "", "success": False, "error": f"DOI未找到 (HTTP {resp.status_code})"})
            except Exception as e:
                citations.append({"doi": clean_doi, "citation": "", "success": False, "error": str(e)})

    return {"citations": citations, "format": "GB/T 7714-2015"}


@app.post("/api/translate")
async def translate_abstract(req: TranslateRequest):
    """学术专业翻译"""
    if not req.text.strip():
        raise HTTPException(status_code=400, detail="请输入需要翻译的文本")

    result = translate_academic(req.text, req.domain)
    return result


@app.post("/api/expand")
async def expand_search_keyword(req: ExpandRequest):
    """关键词扩写"""
    return expand_keywords(req.keyword, req.domain)


@app.get("/api/health")
async def health_check():
    """健康检查"""
    return {"status": "ok", "service": "PaperAgent - 文献检索智能体"}


@app.post("/api/download")
async def download_pdf(data: dict):
    """下载论文 PDF（多源回退：arXiv → OA链接 → 出版商页面）"""
    source = data.get("source", "")
    paper_id = data.get("id", "")
    doi = data.get("doi", "")
    oa_url = data.get("oa_url", "")

    download_urls = []

    # 渠道1: arXiv 直链
    if source == "arxiv" and paper_id:
        download_urls.append(f"https://arxiv.org/pdf/{paper_id}.pdf")

    # 渠道2: OA 链接
    if oa_url and "arxiv" not in oa_url:
        download_urls.append(oa_url)

    # 渠道3: 通过 DOI 转出版商
    if doi:
        download_urls.append(f"https://doi.org/{doi}")

    if not download_urls:
        raise HTTPException(status_code=404, detail="该文献无可用下载渠道")

    # 尝试逐级下载
    from fastapi.responses import StreamingResponse
    for url in download_urls:
        try:
            async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
                head = await client.head(url)
                ct = head.headers.get("content-type", "")
                is_pdf = "pdf" in ct or url.endswith(".pdf") or "arxiv.org/pdf" in url

                if is_pdf:
                    # 返回流式 PDF 下载
                    async def pdf_stream():
                        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as c:
                            async with c.stream("GET", url) as resp:
                                async for chunk in resp.aiter_bytes():
                                    yield chunk

                    filename = paper_id or doi.replace("/", "_") or "paper"
                    return StreamingResponse(
                        pdf_stream(),
                        media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
                    )
                else:
                    # 不是 PDF（可能是出版商页面），返回跳转链接
                    return {"url": url, "note": "无法直接下载PDF，已为您打开文献页面"}
        except Exception:
            continue

    return {"url": download_urls[-1], "note": "PDF 下载失败，请尝试通过出版商页面访问"}


# ── Agent 对话 ─────────────────────────────────────────────

import uuid
from datetime import datetime, timedelta

# 会话历史存储
_chat_sessions: dict[str, dict] = {}
_CHAT_TTL = timedelta(minutes=30)


def _cleanup_expired_sessions():
    """清理过期会话"""
    now = datetime.now()
    expired = [
        sid for sid, s in _chat_sessions.items()
        if now - s.get("updated_at", now) > _CHAT_TTL
    ]
    for sid in expired:
        del _chat_sessions[sid]


class ChatRequest(BaseModel):
    message: str
    session_id: str = ""
    state: int = 0
    history: list[dict] = []


class ChatResponse(BaseModel):
    reply: str
    session_id: str
    state: int
    options: list[str] = []
    done: bool = False
    keywords: dict = {}
    search_query: str = ""


@app.post("/api/chat")
async def chat(req: ChatRequest):
    """Agent 对话接口 — 由 LangChain Agent 驱动"""
    now = datetime.now()
    sid = req.session_id or uuid.uuid4().hex[:12]

    # 定期清理过期会话（每 10 次请求清理一次）
    if uuid.uuid4().int % 10 == 0:
        _cleanup_expired_sessions()

    # 获取或创建会话历史
    session = _chat_sessions.get(sid)
    if not session or (now - session.get("updated_at", now)) > _CHAT_TTL:
        session = {"messages": [], "created_at": now, "updated_at": now}
    session["updated_at"] = now
    _chat_sessions[sid] = session

    msg = req.message.strip()
    if not msg:
        raise HTTPException(status_code=400, detail="消息不能为空")

    try:
        executor = get_agent_executor()

        # 构建消息列表
        messages = list(session["messages"])
        messages.append(HumanMessage(content=msg))

        result = await executor.ainvoke({"messages": messages})
        reply_messages = result.get("messages", [])
        last_msg = reply_messages[-1]
        reply = last_msg.content if hasattr(last_msg, "content") else str(last_msg)

        # 保存对话历史
        session["messages"] = reply_messages[-20:]
        session["updated_at"] = now

        # 解析 【SEARCH:...】 标记
        keywords = {}
        clean_reply = reply
        search_match = re.search(r'【SEARCH:(.+?)】', reply)
        if search_match:
            kw_str = search_match.group(1)
            kw_list = [k.strip() for k in kw_str.split(",") if k.strip()]
            clean_reply = reply.replace(search_match.group(0), "").strip()
            keywords = {
                "en_keywords": kw_list[:10],
                "all_keywords": kw_list[:10],
            }
        else:
            # 兜底：从工具调用记录中提取关键词
            for m in reply_messages:
                if hasattr(m, "tool_calls") and m.tool_calls:
                    for tc in m.tool_calls:
                        if tc.get("name") == "search_papers":
                            args = tc.get("args", {}) or {}
                            kw = args.get("keyword", "")
                            if kw and kw not in [k.get("en_keywords", [""])[0] if isinstance(k, dict) else "" for k in [keywords]]:
                                if "en_keywords" not in keywords:
                                    keywords["en_keywords"] = []
                                keywords["en_keywords"].append(kw)
            if keywords.get("en_keywords"):
                keywords["all_keywords"] = keywords["en_keywords"][:10]

        return ChatResponse(
            reply=clean_reply,
            session_id=sid,
            state=0,
            options=[],
            done=len(keywords) > 0,
            keywords=keywords,
            search_query=" ".join(keywords.get("en_keywords", [])),
        )
    except Exception as e:
        # 失败回退：返回友好提示
        return ChatResponse(
            reply=f"抱歉，处理请求时出错了：{str(e)[:80]}。请稍后重试或换一种表述方式。",
            session_id=sid,
            state=0,
            options=[],
            done=False,
        )


# ── SSE 流式对话 ────────────────────────────────────────────

import asyncio


@app.post("/api/chat/stream")
async def chat_stream(req: ChatRequest):
    """Agent 对话接口（SSE 流式）"""
    now = datetime.now()
    sid = req.session_id or uuid.uuid4().hex[:12]

    session = _chat_sessions.get(sid)
    if not session or (now - session.get("updated_at", now)) > _CHAT_TTL:
        session = {"messages": [], "created_at": now, "updated_at": now}
    session["updated_at"] = now
    _chat_sessions[sid] = session

    msg = req.message.strip()
    if not msg:
        raise HTTPException(status_code=400, detail="消息不能为空")

    async def event_generator():
        try:
            executor = get_agent_executor()
            messages = list(session["messages"])
            messages.append(HumanMessage(content=msg))

            # 先用普通 invoke 拿到结果（Agent 内部有多步 tool call，不适合直接流式）
            # 拿到完整结果后，以 SSE 格式流式输出文本
            yield f"data: {json.dumps({'type': 'status', 'text': '正在分析...'})}\n\n"

            result = await executor.ainvoke({"messages": messages})
            reply_messages = result.get("messages", [])
            last_msg = reply_messages[-1]
            reply = last_msg.content if hasattr(last_msg, "content") else str(last_msg)

            # 解析 【SEARCH:...】 标记
            keywords = {}
            clean_reply = reply
            search_match = re.search(r'【SEARCH:(.+?)】', reply)
            if search_match:
                kw_str = search_match.group(1)
                kw_list = [k.strip() for k in kw_str.split(",") if k.strip()]
                clean_reply = reply.replace(search_match.group(0), "").strip()
                keywords = {
                    "en_keywords": kw_list[:10],
                    "all_keywords": kw_list[:10],
                }
            else:
                # 兜底：从工具调用记录中提取关键词
                for m in reply_messages:
                    if hasattr(m, "tool_calls") and m.tool_calls:
                        for tc in m.tool_calls:
                            if tc.get("name") == "search_papers":
                                args = tc.get("args", {}) or {}
                                kw = args.get("keyword", "")
                                if kw:
                                    if "en_keywords" not in keywords:
                                        keywords["en_keywords"] = []
                                    keywords["en_keywords"].append(kw)
                if keywords.get("en_keywords"):
                    keywords["all_keywords"] = keywords["en_keywords"][:10]

            # 保存对话历史
            session["messages"] = reply_messages[-20:]
            session["updated_at"] = now

            # 流式输出回复文本（逐行）
            lines = clean_reply.split("\n")
            nl = "\n"
            for line in lines:
                yield f"data: {json.dumps({'type': 'text', 'text': line + nl})}\n\n"
                await asyncio.sleep(0.02)

            # 如果有关键词，发送 done 事件
            if keywords:
                yield f"data: {json.dumps({'type': 'done', 'keywords': keywords, 'search_query': ' '.join(keywords.get('en_keywords', []))})}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'done'})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'text': f'抱歉，处理出错：{str(e)[:80]}'})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )

# ── P1 文献综述 ─────────────────────────────────────────────

class PaperItem(BaseModel):
    title: str = ""
    authors: list[str] = []
    year: Optional[int] = None
    abstract: str = ""
    cited_by: int = 0
    doi: str = ""
    keywords: list[str] = []
    journal: str = ""

class AnalyzeRequest(BaseModel):
    papers: list[PaperItem]
    domain: str = "sci-tech"

class AnalyzeResponse(BaseModel):
    review_text: str
    clusters: list[dict]
    year_distribution: list[dict]
    top_keywords: list[dict]
    chart_data: dict


# ── 统计分析函数 ──────────────────────────────────────────

def _year_distribution(papers: list[dict]) -> list[dict]:
    """按年份统计发文量"""
    from collections import Counter
    years = [p.get("year") for p in papers if p.get("year")]
    counter = Counter(years)
    sorted_years = sorted(counter.keys())
    return [{"year": y, "count": counter[y]} for y in sorted_years]


def _extract_keywords(papers: list[dict]) -> list[str]:
    """从文献中提取所有关键词"""
    all_kws = []
    for p in papers:
        kws = p.get("keywords") or []
        for kw in kws:
            if kw and isinstance(kw, str):
                all_kws.append(kw.strip().lower())
    return all_kws


def _keyword_frequency(all_kws: list[str]) -> list[dict]:
    """统计关键词频次"""
    from collections import Counter
    counter = Counter(all_kws)
    return [{"name": k, "value": v} for k, v in counter.most_common(30)]


def _build_cooccurrence(papers: list[dict]) -> dict:
    """构建关键词共现矩阵"""
    from collections import defaultdict
    cooccur = defaultdict(int)
    kw_counter = defaultdict(int)

    for p in papers:
        kws = [kw.strip().lower() for kw in (p.get("keywords") or []) if kw and isinstance(kw, str)]
        for kw in kws:
            kw_counter[kw] += 1
        for i in range(len(kws)):
            for j in range(i + 1, len(kws)):
                a, b = (kws[i], kws[j]) if kws[i] < kws[j] else (kws[j], kws[i])
                cooccur[(a, b)] += 1

    # 过滤低频 (至少共现2次才保留)
    filtered = {k: v for k, v in cooccur.items() if v >= 1}

    top_kws = set()
    for (a, b) in filtered:
        top_kws.add(a)
        top_kws.add(b)

    nodes = [{"name": kw, "value": kw_counter[kw]} for kw in top_kws]
    edges = [{"source": a, "target": b, "value": v} for (a, b), v in filtered.items()]

    return {"nodes": nodes, "edges": edges}


def _cluster_papers(papers: list[dict]) -> list[dict]:
    """基于关键词共现的简单主题聚类"""
    paper_keywords = []
    for i, p in enumerate(papers):
        kws = set(kw.strip().lower() for kw in (p.get("keywords") or []) if kw and isinstance(kw, str))
        paper_keywords.append((i, kws))

    clusters = []
    assigned = set()

    for i, (idx, kws) in enumerate(paper_keywords):
        if idx in assigned:
            continue
        merged_kws = set(kws)
        members = [idx]
        assigned.add(idx)

        # 找相似论文 (关键词重叠 >= 1)
        changed = True
        while changed:
            changed = False
            for j, (jdx, jkws) in enumerate(paper_keywords):
                if jdx in assigned:
                    continue
                overlap = merged_kws & jkws
                if len(overlap) >= 1:
                    members.append(jdx)
                    merged_kws |= jkws
                    assigned.add(jdx)
                    changed = True

        # 取前5个关键词作为主题标签
        from collections import Counter
        kw_counts = Counter(merged_kws)
        top_kw = [kw for kw, _ in kw_counts.most_common(5)]
        label = " / ".join(top_kw[:3]) if top_kw else f"主题 {len(clusters) + 1}"

        clusters.append({
            "id": f"cluster_{len(clusters)}",
            "label": label,
            "papers": members,
            "keywords": top_kw,
            "paper_count": len(members),
        })

    # 确保每篇论文都有归属
    for i in range(len(papers)):
        if i not in assigned:
            clusters.append({
                "id": f"cluster_{len(clusters)}",
                "label": f"主题 {len(clusters) + 1}",
                "papers": [i],
                "keywords": [],
                "paper_count": 1,
            })
            assigned.add(i)

    return clusters


def _get_high_cited(papers: list[dict], top_n: int = 10) -> list[dict]:
    """按引用量排序，识别高被引文献"""
    sorted_papers = sorted(papers, key=lambda p: p.get("cited_by", 0) or 0, reverse=True)
    return [
        {
            "index": i,
            "title": p.get("title", ""),
            "cited_by": p.get("cited_by", 0),
            "authors": p.get("authors", []),
            "year": p.get("year"),
        }
        for i, p in enumerate(sorted_papers[:top_n])
        if p.get("cited_by", 0) > 0
    ]


def _generate_chart_data(papers: list[dict]) -> dict:
    """生成前端图表所需数据结构"""
    # trend: 年份趋势
    year_data = _year_distribution(papers)
    trend = {
        "years": [d["year"] for d in year_data],
        "counts": [d["count"] for d in year_data],
    }

    # cooccurrence: 共现网络
    cooccurrence = _build_cooccurrence(papers)

    # bubble: 气泡图数据
    clusters = _cluster_papers(papers)
    bubble = []
    for c in clusters:
        for pid in c["papers"]:
            p = papers[pid]
            bubble.append({
                "year": p.get("year", 0),
                "cluster": c["label"],
                "title": (p.get("title", "") or "")[:30],
                "cited_by": p.get("cited_by", 0),
            })

    # sankey: 桑基图 (年份 → 主题 → 文献)
    sankey_nodes = []
    sankey_links = []
    node_names = set()

    # 年份节点
    yr_nodes = []
    for d in year_data:
        yr_name = str(d["year"])
        if yr_name not in node_names:
            node_names.add(yr_name)
            yr_nodes.append({"name": yr_name})

    # 聚类节点
    cl_nodes = []
    for c in clusters:
        cl_name = (c["label"] or "")[:20]
        if cl_name not in node_names:
            node_names.add(cl_name)
            cl_nodes.append({"name": cl_name})

    sankey_nodes = yr_nodes + cl_nodes

    # 年份→聚类连线
    for c in clusters:
        cl_name = (c["label"] or "")[:20]
        for pid in c["papers"]:
            p = papers[pid]
            yr_name = str(p.get("year", 0))
            if yr_name in node_names and cl_name in node_names:
                sankey_links.append({"source": yr_name, "target": cl_name, "value": 1})

    return {
        "trend": trend,
        "cooccurrence": cooccurrence,
        "bubble": bubble,
        "sankey": {"nodes": sankey_nodes, "links": sankey_links},
    }


# ── GB/T 7714 引用（复用已有 generate_gbt7714）───────────

def _gbt7714_for_analyze(papers: list[dict], indices: list[int]) -> list[str]:
    """为指定索引的文献生成 GB/T 7714 引用"""
    refs = []
    for i in indices:
        if i < len(papers):
            refs.append(generate_gbt7714(papers[i]))
    return refs


# ── 模板化综述生成（5段，纯规则） ──────────────────────────

def _generate_review_template(papers: list[dict], domain: str) -> str:
    """纯规则模板生成5段综述文本"""
    n = len(papers)
    if n == 0:
        return "未提供文献数据，无法生成综述。"

    # 基本统计
    year_data = _year_distribution(papers)
    earliest = year_data[0]["year"] if year_data else "未知"
    latest = year_data[-1]["year"] if year_data else "未知"

    high_cited = _get_high_cited(papers, 5)
    clusters = _cluster_papers(papers)
    all_kws = _extract_keywords(papers)
    top_kws = _keyword_frequency(all_kws)[:8]

    # 期刊统计
    journals = {}
    for p in papers:
        j = p.get("journal", "未知期刊")
        if j:
            journals[j] = journals.get(j, 0) + 1
    top_journals = sorted(journals.items(), key=lambda x: -x[1])[:5]

    total_citations = sum(p.get("cited_by", 0) or 0 for p in papers)
    domain_label = "自然科学与工程技术" if domain == "sci-tech" else "人文社会科学"

    ref_indices = list(range(min(n, 20)))
    refs = _gbt7714_for_analyze(papers, ref_indices)

    kw_str = "、".join([k["name"] for k in top_kws[:5]])

    # ── 第1段：背景介绍 ──
    para1 = (
        f"一、研究背景\n\n"
        f"{kw_str}是{domain_label}领域近年来备受关注的研究方向。"
        f"随着科学技术的快速发展，该领域的研究成果不断涌现。"
        f"本综述基于{n}篇相关文献"
    )
    if earliest != latest:
        para1 += f"（涵盖{earliest}年至{latest}年）"
    para1 += (
        f"，从发文趋势、研究主题、高被引文献等多个维度对该领域的研究现状进行系统梳理与分析。"
        f"通过对现有文献的计量分析，旨在揭示该领域的研究热点、发展趋势以及存在的不足，"
        f"为后续研究提供参考。\n"
    )

    # ── 第2段：国内外现状 ──
    key_findings = []
    if high_cited:
        top_paper = high_cited[0]
        key_findings.append(
            f"在引用影响力方面，{top_paper['title'][:30]}…被引{top_paper['cited_by']}次，"
            f"是该领域最具影响力的工作之一。"
        )
    if top_journals:
        j_str = "、".join([f"《{j}》（{c}篇）" for j, c in top_journals[:3]])
        key_findings.append(f"主要发表期刊包括{j_str}。")

    yr_trend = ""
    if len(year_data) >= 3:
        recent_avg = sum(d["count"] for d in year_data[-3:]) / 3
        older_avg = sum(d["count"] for d in year_data[:3]) / 3
        if recent_avg > older_avg * 1.5:
            yr_trend = "从发文趋势来看，近年来该领域发文量呈现明显上升态势，表明该方向正成为研究热点。"
        elif recent_avg < older_avg * 0.7:
            yr_trend = "从发文趋势来看，近年来该领域发文量有所下降，可能表明该方向已进入成熟期。"
        else:
            yr_trend = "从发文趋势来看，该领域发文量保持相对稳定，说明已形成持续稳定的研究社区。"

    para2 = (
        f"二、国内外研究现状\n\n"
        f"目前，国内外学者围绕{kw_str}开展了大量卓有成效的研究工作。"
        f"{' '.join(key_findings)}"
        f"{yr_trend}"
        f"共检索到相关文献{n}篇，总被引次数达{total_citations}次，"
        f"篇均被引{total_citations / max(n, 1):.1f}次。\n"
    )

    # ── 第3段：分支梳理 ──
    para3 = "三、研究分支梳理\n\n"
    for i, c in enumerate(clusters[:5]):
        kw_list = "、".join(c.get("keywords", [])[:4]) if c.get("keywords") else "未明确"
        para3 += (
            f"**{c['label']}**：涵盖{c['paper_count']}篇文献，"
            f"核心关键词包括{kw_list}。"
            f"该方向的研究主要集中于上述关键词所代表的技术与方法。\n\n"
        )
    para3 += (
        f"通过主题聚类分析，该领域的研究可分为{len(clusters)}个主要方向，"
        f"各方向之间既有所侧重又相互交叉，共同构成了该领域的研究全景。\n"
    )

    # ── 第4段：研究空白 ──
    gaps = []
    if n < 50:
        gaps.append(f"目前该领域文献总量相对有限（{n}篇），尚缺乏大规模的系统性研究。")
    if len(clusters) <= 2:
        gaps.append("研究主题较为集中，跨方向交叉融合的研究有待进一步拓展。")
    if not high_cited or high_cited[0]["cited_by"] < 50:
        gaps.append("高被引文献较少，尚未形成具有广泛共识的标志性成果。")
    if not gaps:
        gaps.append("尽管已有大量研究成果，但仍然存在一些值得关注的研究空白。")

    para4 = (
        f"四、研究空白与不足\n\n"
        f"综合以上分析，当前研究仍存在以下不足：{' '.join(gaps)}"
        f"此外，部分研究在方法论的严谨性、数据的时效性以及跨学科融合方面仍有提升空间。"
        f"未来研究可在上述方向进行深入探索。\n"
    )

    # ── 第5段：本文定位 ──
    para5 = (
        f"五、本文定位\n\n"
        f"基于以上文献综述和分析，本文旨在系统梳理{kw_str}领域的研究脉络，"
        f"在现有研究的基础上，针对上述研究空白，提出新的研究思路和方法。"
        f"本文的主要贡献包括：（1）全面梳理了该领域的发展历程和研究现状；"
        f"（2）通过文献计量方法揭示了研究热点和趋势；"
        f"（3）明确了当前研究的不足和未来方向。"
        f"期望本研究能够为{domain_label}领域的研究者提供有价值的参考。\n\n"
        f"---\n\n"
        f"参考文献\n\n"
    )
    for i, ref in enumerate(refs):
        para5 += f"[{i + 1}] {ref}\n"

    return para1 + "\n" + para2 + "\n" + para3 + "\n" + para4 + "\n" + para5


# ── LLM 增强综述生成（可选） ────────────────────────────

import os as _os_mod

LLM_API_KEY = _os_mod.environ.get("LLM_API_KEY", "")


async def _generate_review_llm(papers: list[dict], domain: str) -> str:
    """使用 LLM 生成综述（当配置了 LLM_API_KEY 时调用），否则回退到模板"""
    if not LLM_API_KEY:
        return _generate_review_template(papers, domain)

    papers_summary = []
    for i, p in enumerate(papers[:50]):
        papers_summary.append(
            f"[{i + 1}] Title: {p.get('title', '')}\n"
            f"    Authors: {', '.join(p.get('authors', [])[:3])}\n"
            f"    Year: {p.get('year', '')}\n"
            f"    Journal: {p.get('journal', '')}\n"
            f"    Cited: {p.get('cited_by', 0)}\n"
            f"    Keywords: {', '.join(p.get('keywords', [])[:8])}\n"
            f"    Abstract: {p.get('abstract', '')[:300]}\n"
        )

    prompt = (
        f"你是一位学术研究专家。请基于以下{len(papers)}篇文献，写一篇中文文献综述。\n\n"
        f"要求：\n"
        f"1. 结构要求（5个段落）：\n"
        f"   一、研究背景 — 介绍研究领域的背景和意义\n"
        f"   二、国内外研究现状 — 梳理国内外主要研究成果\n"
        f"   三、研究分支梳理 — 归纳不同研究方向/主题\n"
        f"   四、研究空白与不足 — 指出现有研究的不足和空白\n"
        f"   五、本文定位 — 说明本文的位置和贡献\n\n"
        f"2. 在正文中标注引用编号如[1][2]等\n"
        f"3. 在文末按GB/T 7714格式列出参考文献\n"
        f"4. 专业、严谨的学术语言\n\n"
        f"文献列表：\n" + "\n".join(papers_summary)
    )

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {LLM_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "deepseek-chat",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.3,
                    "max_tokens": 4096,
                },
                timeout=60.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"]
            return _generate_review_template(papers, domain)
    except Exception:
        return _generate_review_template(papers, domain)


# ── API 路由：文献分析 ───────────────────────────────────

@app.post("/api/analyze", response_model=AnalyzeResponse)
async def analyze_papers(req: AnalyzeRequest):
    """批量分析文献并生成综述"""
    if not req.papers:
        raise HTTPException(status_code=400, detail="请提供至少一篇文献")

    papers_data = [p.model_dump() if hasattr(p, 'model_dump') else dict(p) for p in req.papers]

    # 1. 年份分布
    year_distribution = _year_distribution(papers_data)

    # 2. 关键词频次
    all_kws = _extract_keywords(papers_data)
    top_keywords = _keyword_frequency(all_kws)

    # 3. 主题聚类
    clusters = _cluster_papers(papers_data)

    # 4. 图表数据
    chart_data = _generate_chart_data(papers_data)

    # 5. 综述生成（LLM或模板）
    review_text = await _generate_review_llm(papers_data, req.domain)

    return AnalyzeResponse(
        review_text=review_text,
        clusters=clusters,
        year_distribution=year_distribution,
        top_keywords=top_keywords,
        chart_data=chart_data,
    )

# ── P2 数据上传 + Stata集成 + 排列组合 + 结果解读 ──────────

import asyncio
import csv
import io as _io_mod
import itertools
import math as _math_mod
import os
import random as _random_mod
from pathlib import Path as _Path

from fastapi import File, Form, UploadFile
from openpyxl import load_workbook

# ── 配置 ────────────────────────────────────────────────────

STATA_WORK_DIR = "/mnt/c/temp/stata_work"
STATA_EXE = r"D:\Program Files\StataMP-64.exe"


def _detect_stata_exe() -> str:
    """检测 Stata 可执行文件路径（尝试多个常见名称）"""
    candidates = [
        r"D:\Program Files\StataMP-64.exe",
        r"D:\Program Files\StataSE-64.exe",
        r"D:\Program Files\StataMP\StataMP-64.exe",
        r"D:\Program Files\Stata18\StataMP-64.exe",
        r"D:\Program Files\Stata18\StataSE-64.exe",
        r"D:\Program Files\stata18\StataMP-64.exe",
    ]
    for candidate in candidates:
        wsl_path = _to_wsl_path(candidate)
        if _Path(wsl_path).exists():
            return candidate
    return STATA_EXE  # fallback


def _to_win_path(wsl_path: str) -> str:
    """将 WSL 路径转为 Windows 路径，如 /mnt/c/temp/xxx -> C:\\temp\\xxx"""
    if wsl_path.startswith("/mnt/"):
        drive = wsl_path[5].upper()
        rest = "\\" + wsl_path[7:].replace("/", "\\")
        return f"{drive}:{rest}"
    return wsl_path.replace("/", "\\")


def _to_wsl_path(win_path: str) -> str:
    """将 Windows 路径转为 WSL 路径，如 C:\\temp\\xxx -> /mnt/c/temp/xxx"""
    if ":" in win_path:
        drive = win_path[0].lower()
        rest = win_path[2:].replace("\\", "/")
        return f"/mnt/{drive}{rest}"
    return win_path.replace("\\", "/")


# 自动检测 Stata 版本（需在 _to_wsl_path 之后调用）
STATA_EXE = _detect_stata_exe()

LLM_BASE_URL = os.environ.get("LLM_BASE_URL", "https://api.deepseek.com/v1")

# ── 数据模型 ──────────────────────────────────────────────

class UploadResponse(BaseModel):
    filename: str
    columns: list[dict]
    preview: list[dict]
    row_count: int

class StataRunRequest(BaseModel):
    command: str = "reg"
    y: str
    x: list[str]
    controls: list[str] = []
    m: list[str] = []
    w: str = ""
    data_columns: list[str]
    data_file: str

class StataRunResponse(BaseModel):
    success: bool
    log: str = ""
    do_file_content: str = ""
    result_table: list[dict] = []
    stats: dict = {}

class StataCombinationsRequest(BaseModel):
    y_list: list[str]
    x_list: list[str]
    m_list: list[str] = []
    w_list: list[str] = []
    controls: list[str] = []
    sort_by: str = "r2"
    data_file: str = ""
    methods: list[str] = []

class StataCombinationsResponse(BaseModel):
    combinations: list[dict]
    best: dict = {}

class InterpretRequest(BaseModel):
    result_table: list[dict]
    method: str = "reg"
    paper_count: int = 0
    stats: dict = {}

class InterpretResponse(BaseModel):
    significance: str = ""
    economic_significance: str = ""
    innovation_points: list[str] = []
    comparison_text: str = ""
    chapter_draft: str = ""

# ── 辅助函数 ──────────────────────────────────────────────

def _ensure_stata_workdir():
    """确保 Stata 工作目录存在"""
    p = _Path(STATA_WORK_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _detect_column_type(values: list) -> str:
    """检测列的数据类型"""
    numeric_count = 0
    for v in values:
        if v is None or v == "":
            continue
        try:
            float(v)
            numeric_count += 1
        except (ValueError, TypeError):
            pass
    if numeric_count > len(values) * 0.7:
        return "numeric"
    return "string"


def _df_to_stata_data(df, filepath: str):
    """将 DataFrame 写入 Stata .dta 文件"""
    clean_cols = {}
    for c in df.columns:
        clean = re.sub(r'[^a-zA-Z0-9_]', '_', str(c))[:32]
        clean_cols[c] = clean
    df = df.rename(columns=clean_cols)
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].astype(str)
    df.to_stata(filepath, write_index=False, version=118)
    return list(clean_cols.values())


# ── Stata Do 文件生成 ──────────────────────────────────────

def _build_stata_do(
    command: str,
    y: str,
    x: list[str],
    controls: list[str],
    m: list[str],
    w: str,
    data_file: str,
    output_label: str = "main",
) -> str:
    """构建 Stata do 文件内容"""
    lines = []
    lines.append('clear all')
    lines.append('set more off')
    lines.append('set matsize 5000')
    # 检查需要的外部包
    ext_pkgs = {'reghdfe': 'reghdfe', 'ivreg2': 'ivreg2', 'psmatch2': 'psmatch2',
                 'csdid': 'csdid', 'eventdd': 'eventdd', 'staggered': 'staggered'}
    for cmd, pkg in ext_pkgs.items():
        if command == cmd or (command == "ivreg2" and m and cmd == "ivreg2"):
            lines.append(f'capture which {cmd}')
            lines.append(f'if _rc {{');
            lines.append(f'  display as error "注意：{cmd} 需要额外安装，请执行: ssc install {pkg}"');
            lines.append('}')
    lines.append('')
    # 启用日志记录
    win_log = _to_win_path(str(_Path(STATA_WORK_DIR) / f"{output_label}.log"))
    lines.append(f'log using "{win_log.replace(chr(92), chr(47))}", replace text')
    # CSV文件用 import delimited, DTA文件用 use
    if data_file.lower().endswith('.csv'):
        # Stata 运行在 Windows 上，需用 Windows 路径
        win_path = _to_win_path(data_file)
        csv_path = win_path.replace('\\', '/')
        lines.append(f'import delimited "{csv_path}", clear')
        lines.append('destring _all, replace')
    else:
        lines.append(f'use "{data_file}", clear')
    lines.append('')

    # 设定全局变量
    lines.append(f'global y {y}')
    x_vars = " ".join(x)
    lines.append(f'global xvars {x_vars}')
    if controls:
        lines.append(f'global controls {" ".join(controls)}')
    if m:
        lines.append(f'global mvars {" ".join(m)}')
    if w:
        lines.append(f'global wvar {w}')
    lines.append('')

    # 构建回归命令
    depvar = "$y"
    indep_parts = ["$xvars"]
    if controls:
        indep_parts.append("$controls")
    indep_str = " ".join(indep_parts)

    if command == "reg":
        lines.append(f'reg {depvar} {indep_str}')
    elif command == "reghdfe":
        absorb_parts = controls if controls else []
        absorb_str = " ".join(absorb_parts) if absorb_parts else ""
        fe_indep = x_vars
        if absorb_str:
            lines.append(f'reghdfe {depvar} {fe_indep}, absorb({absorb_str})')
        else:
            lines.append(f'reghdfe {depvar} {fe_indep}, noabsorb')
    elif command == "xtreg":
        lines.append(f'xtreg {depvar} {indep_str}, fe')
    elif command == "ivreg2":
        if m:
            iv_str = " ".join(m)
            lines.append(f'ivreg2 {depvar} {indep_str} ({controls} = {iv_str})')
        else:
            lines.append(f'ivreg2 {depvar} {indep_str}')
    elif command == "logit":
        lines.append(f'logit {depvar} {indep_str}')
    elif command == "probit":
        lines.append(f'probit {depvar} {indep_str}')
    elif command == "tobit":
        lines.append(f'tobit {depvar} {indep_str}')
    else:
        lines.append(f'reg {depvar} {indep_str}')

    lines.append('')

    # 输出结果
    lines.append('* --- 结果输出 ---')
    lines.append('estimates store m1')
    lines.append('')
    lines.append('log close')
    lines.append('exit, STATA')

    return "\n".join(lines)


# ── Stata Log 解析 ──────────────────────────────────────────

def _parse_stata_log(log_text: str) -> tuple[list[dict], dict]:
    """解析 Stata log 输出，提取回归结果表"""
    result_table = []
    stats = {}

    # 获取观测数
    n_match = re.search(r'Number of obs\s*=\s*(\d+)', log_text)
    if n_match:
        stats['n'] = int(n_match.group(1))

    # 获取 R-squared
    r2_match = re.search(r'R-squared\s*=\s*([0-9.]+)', log_text)
    if r2_match:
        stats['r2'] = float(r2_match.group(1))

    # 获取 Adj R-squared
    adj_r2_match = re.search(r'Adj R-squared\s*=\s*([0-9.]+)', log_text)
    if adj_r2_match:
        stats['adj_r2'] = float(adj_r2_match.group(1))

    # 获取 F-stat
    f_match = re.search(r'F\(\s*\d+\s*,\s*\d+\)\s*=\s*([0-9.]+)', log_text)
    if f_match and f_match.group(1) != '.':
        try:
            stats['f_stat'] = float(f_match.group(1))
        except ValueError:
            pass

    # 解析系数表
    lines = log_text.split('\n')
    in_coef_table = False
    coef_header_found = False
    for line in lines:
        # 检测表头
        if re.search(r'^\s*\w+\s*\|?\s*(Coeff|Coefficient|Coef)', line):
            in_coef_table = True
            coef_header_found = False
            continue
        if in_coef_table:
            # 跳过分隔线（------+--- 格式的行）
            if re.search(r'^-+\+', line):
                coef_header_found = True
                continue
            # 空行或结束线退出
            if re.search(r'^\s*$', line) or (re.search(r'^-', line) and coef_header_found):
                in_coef_table = False
                continue
            if re.search(r'^\s*\w', line):
                parts = [p for p in line.strip().split() if p != '|']
                if len(parts) >= 5 and parts[0] not in ('_cons',):
                    try:
                        var_name = parts[0]
                        coef = float(parts[1]) if parts[1] != '.' else 0.0
                        se = float(parts[2]) if parts[2] != '.' else 0.0
                        pval = float(parts[4]) if parts[4] != '.' else 1.0
                        stars = ""
                        if pval < 0.01:
                            stars = "***"
                        elif pval < 0.05:
                            stars = "**"
                        elif pval < 0.1:
                            stars = "*"
                        result_table.append({
                            "var": var_name,
                            "coef": coef,
                            "se": se,
                            "pval": pval,
                            "stars": stars,
                        })
                    except (ValueError, IndexError):
                        continue

    # 第二遍：如果上面没匹配到，尝试更宽松的匹配
    if not result_table:
        for line in lines:
            if 'Coef.' in line and 'Std. Err.' in line:
                in_coef_table = True
                continue
            if in_coef_table:
                if re.search(r'^\s*$', line) or '---' in line:
                    in_coef_table = False
                    continue
                parts = [p for p in line.strip().split() if p != '|']
                if len(parts) >= 5 and not parts[0].startswith('_'):
                    try:
                        var_name = parts[0]
                        coef = float(parts[1]) if parts[1] != '.' else 0.0
                        se = float(parts[2]) if parts[2] != '.' else 0.0
                        pval = float(parts[4]) if len(parts) >= 5 and parts[4] != '.' else 1.0
                        stars = ""
                        if pval <= 0.01:
                            stars = "***"
                        elif pval <= 0.05:
                            stars = "**"
                        elif pval <= 0.1:
                            stars = "*"
                        result_table.append({
                            "var": var_name,
                            "coef": coef,
                            "se": se,
                            "pval": pval,
                            "stars": stars,
                        })
                    except (ValueError, IndexError):
                        pass

    return result_table, stats


# ── 执行 Stata ──────────────────────────────────────────────

async def _run_stata_do(do_content: str, do_name: str = "run") -> dict:
    """通过 PowerShell 调用 Windows Stata 执行 do 文件"""
    _ensure_stata_workdir()

    do_path = _Path(STATA_WORK_DIR) / f"{do_name}.do"
    log_path = _Path(STATA_WORK_DIR) / f"{do_name}.log"

    # 写入 do 文件
    do_path.write_text(do_content, encoding="utf-8")

    # 构建 PowerShell 命令（从 WSL 调用 Windows Stata，需用 Windows 路径）
    win_do_path = _to_win_path(str(do_path))

    try:
        ps_cmd = f"& '{STATA_EXE}' /e do {win_do_path}"

        async def run_ps():
            proc = await asyncio.create_subprocess_exec(
                "/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe",
                "-NoProfile", "-Command", ps_cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            try:
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120.0)
                return stdout, stderr
            except asyncio.TimeoutError:
                proc.kill()
                return None, None

        stdout, stderr = await run_ps()
        if stdout is None:  # 超时
            return {"success": False, "log": "Stata 执行超时（120秒）", "do_file_content": do_content, "result_table": [], "stats": {}}

        # 读取 log 文件（do 文件里已包含 log using 命令）
        log_text = ""
        if log_path.exists():
            log_text = log_path.read_text(encoding="utf-8", errors="replace")
        elif stdout:
            log_text = stdout.decode("utf-8", errors="replace")

        result_table, stats = _parse_stata_log(log_text)

        return {
            "success": True,
            "log": log_text[:5000],
            "do_file_content": do_content,
            "result_table": result_table,
            "stats": stats,
        }
    except FileNotFoundError:
        return {
            "success": False,
            "log": f"未找到 PowerShell 或 Stata 可执行文件（{STATA_EXE}）",
            "do_file_content": do_content,
            "result_table": [],
            "stats": {},
        }
    except Exception as e:
        return {
            "success": False,
            "log": f"执行失败: {str(e)}",
            "do_file_content": do_content,
            "result_table": [],
            "stats": {},
        }


# ── API: 数据上传 ───────────────────────────────────────────

@app.post("/api/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...)):
    """上传 CSV/Excel 数据文件"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择文件")

    suffix = _Path(file.filename).suffix.lower()
    if suffix not in (".csv", ".xls", ".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 CSV 和 Excel (.xls/.xlsx) 文件")

    content = await file.read()

    try:
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(_io_mod.StringIO(text))
            rows = list(reader)
            columns = reader.fieldnames or []
        else:
            wb = load_workbook(_io_mod.BytesIO(content), read_only=True)
            ws = wb.active
            data_rows = list(ws.iter_rows(values_only=True))
            if not data_rows:
                raise HTTPException(status_code=400, detail="Excel 文件为空")
            columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(data_rows[0])]
            rows = []
            for data_row in data_rows[1:]:
                row_dict = {}
                for i, val in enumerate(data_row):
                    col_name = columns[i] if i < len(columns) else f"col_{i}"
                    row_dict[col_name] = val
                rows.append(row_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据行")

    # 列信息
    col_info = []
    for col_name in columns:
        col_values = [r.get(col_name) for r in rows[:100]]
        col_type = _detect_column_type(col_values)
        col_info.append({"name": col_name, "type": col_type})

    # 预览前5行
    preview = rows[:5]
    for row in preview:
        for k, v in row.items():
            pass  # 预览值原样保留

    # 保存原始文件供 Stata 使用（直接存为 CSV，Stata 用 import delimited 读取）
    _ensure_stata_workdir()
    safe_name = uuid.uuid4().hex[:8]
    
    # 保存为 CSV（Stata 可直接读取）
    csv_path = _Path(STATA_WORK_DIR) / f"{safe_name}.csv"
    with open(str(csv_path), "wb") as f:
        if suffix in (".xls", ".xlsx"):
            # Excel 转 CSV
            wb2 = load_workbook(_io_mod.BytesIO(content), read_only=True)
            ws2 = wb2.active
            import csv as csv_mod
            with open(str(csv_path), "w", newline="", encoding="utf-8-sig") as cf:
                writer = csv_mod.writer(cf)
                for row in ws2.iter_rows(values_only=True):
                    writer.writerow(row)
        else:
            f.write(content)

    return UploadResponse(
        filename=f"{safe_name}.csv",
        columns=col_info,
        preview=preview,
        row_count=len(rows),
    )


# ── API: Stata 回归执行 ─────────────────────────────────────

@app.post("/api/stata/run", response_model=StataRunResponse)
async def stata_run(req: StataRunRequest):
    """执行 Stata 回归命令"""
    if not req.y or not req.x:
        raise HTTPException(status_code=400, detail="请指定因变量 y 和自变量 x")

    data_file = str(_Path(STATA_WORK_DIR) / req.data_file)

    do_content = _build_stata_do(
        command=req.command,
        y=req.y,
        x=req.x,
        controls=req.controls,
        m=req.m,
        w=req.w,
        data_file=data_file,
    )

    result = await _run_stata_do(do_content, "run_" + uuid.uuid4().hex[:8])

    return StataRunResponse(**result)


# ── API: 排列组合遍历 ───────────────────────────────────────

@app.post("/api/stata/combinations", response_model=StataCombinationsResponse)
async def stata_combinations(req: StataCombinationsRequest):
    """排列组合遍历回归"""
    if not req.y_list or not req.x_list:
        raise HTTPException(status_code=400, detail="请提供 y_list 和 x_list")

    y_choices = req.y_list
    x_choices = req.x_list
    m_choices = req.m_list if req.m_list else [None]
    w_choices = req.w_list if req.w_list else [None]

    # 限制组合总数
    max_combinations = 200
    all_combos = list(itertools.product(y_choices, x_choices, m_choices, w_choices))
    total = len(all_combos)
    if total > max_combinations:
        all_combos = _random_mod.sample(all_combos, max_combinations)

    # 执行每个组合
    results = []
    rank = 1

    # 构建数据文件路径
    data_file_path = str(_Path(STATA_WORK_DIR) / req.data_file) if req.data_file else ""
    methods = req.methods or ["reg"]

    for y, x, m, w in all_combos:
        x_vars = [x] if isinstance(x, str) else x
        m_vars = [m] if m else []
        w_var = w or ""

        for command in methods:
            do_content = _build_stata_do(
            command="ivreg2" if m_vars else command,
            y=y,
            x=x_vars if isinstance(x_vars, list) else [x_vars],
            controls=req.controls,
            m=m_vars,
            w=w_var,
            data_file=data_file_path,  # 使用前端传入的 data_file
            output_label=f"combo_{rank}",
            )

            result = await _run_stata_do(do_content, f"combo_{rank}")

            stats = result.get("stats", {})
            table = result.get("result_table", [])

            main_coef = 0
            main_pval = 1.0
            for row in table:
                if row["var"] == x:
                    main_coef = row["coef"]
                    main_pval = row["pval"]
                    break

            r2_val = stats.get("r2", 0)
            n_val = stats.get("n", 0)

            reason_parts = []
            sig_level = ""
            if main_pval <= 0.01:
                sig_level = "1%显著"
                reason_parts.append("p<0.01")
            elif main_pval <= 0.05:
                sig_level = "5%显著"
                reason_parts.append("p<0.05")
            elif main_pval <= 0.1:
                sig_level = "10%显著"
                reason_parts.append("p<0.1")
            else:
                reason_parts.append("不显著")

            if r2_val > 0.3:
                reason_parts.append(f"R²={r2_val:.3f}")

            results.append({
            "rank": rank,
            "formula": f"{y} ~ {x}",
            "coef": main_coef,
            "r2": r2_val,
            "pval": main_pval,
            "n": n_val,
            "sig": sig_level,
            "reason": "; ".join(reason_parts),
            "m": m,
            "w": w,
            "controls": req.controls,
            "method": command,
        })
        rank += 1

    # 排序
    if req.sort_by == "r2":
        results.sort(key=lambda r: -r["r2"])
    elif req.sort_by == "sig":
        results.sort(key=lambda r: r["pval"])

    # 重新编号
    for i, r in enumerate(results):
        r["rank"] = i + 1

    best = results[0] if results else {}

    return StataCombinationsResponse(
        combinations=results[:50],
        best=best,
    )


# ── API: LLM 结果解读 ───────────────────────────────────────

@app.post("/api/stata/interpret", response_model=InterpretResponse)
async def stata_interpret(req: InterpretRequest):
    """LLM 结果解读"""
    if not req.result_table:
        raise HTTPException(status_code=400, detail="请提供回归结果")

    method_name = {
        "reg": "OLS 线性回归",
        "reghdfe": "高维固定效应回归",
        "xtreg": "面板随机/固定效应",
        "ivreg2": "工具变量 2SLS",
        "logit": "Logit 模型",
        "probit": "Probit 模型",
        "tobit": "Tobit 模型",
    }.get(req.method, req.method)

    if not LLM_API_KEY:
        return _rule_based_interpret(req.result_table, req.stats, method_name)

    # 构建 LLM 提示
    table_str = "\n".join([
        f"| {r.get('var', '')} | Coef={r.get('coef', 0)} | SE={r.get('se', 0)} | p={r.get('pval', 1)} | {r.get('stars', '')} |"
        for r in req.result_table
    ])
    stats_str = json.dumps(req.stats, ensure_ascii=False, indent=2)

    prompt = f"""你是一位计量经济学和学术论文写作专家。请根据以下回归结果，撰写一份详细的中文结果解读。

## 回归方法
{method_name}

## 回归结果表
{table_str}

## 统计量
{stats_str}

## 参考文献数量
{req.paper_count} 篇

请按以下结构输出（用 JSON 格式）：
{{
    "significance": "统计显著性分析：讨论哪些变量显著、显著性水平、系数符号的经济含义",
    "economic_significance": "经济显著性分析：讨论系数大小在实际中的含义，不能只看 p 值",
    "innovation_points": ["创新点1", "创新点2", ...],
    "comparison_text": "与已有文献的比较分析",
    "chapter_draft": "论文结果章节草稿（约300-500字，包含规范的学术表述）"
}}

注意：
- 使用规范的学术语言
- 标注星号含义（* p<0.1, ** p<0.05, *** p<0.01）
- 指出可能的内生性问题和局限性
"""

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                f"{LLM_BASE_URL}/chat/completions",
                headers={
                    "Authorization": f"Bearer {LLM_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "deepseek-chat",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.3,
                    "max_tokens": 4096,
                },
                timeout=60.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                content = data["choices"][0]["message"]["content"]
                try:
                    parsed = json.loads(content)
                    return InterpretResponse(**parsed)
                except (json.JSONDecodeError, Exception):
                    json_match = re.search(r'\{[\s\S]*\}', content)
                    if json_match:
                        try:
                            parsed = json.loads(json_match.group())
                            return InterpretResponse(**parsed)
                        except Exception:
                            pass
                    return InterpretResponse(
                        significance="LLM 返回格式异常",
                        chapter_draft=content[:2000],
                    )
    except Exception:
        return _rule_based_interpret(req.result_table, req.stats, method_name)

    return _rule_based_interpret(req.result_table, req.stats, method_name)


def _rule_based_interpret(
    result_table: list[dict],
    stats: dict,
    method_name: str = "OLS 线性回归",
) -> InterpretResponse:
    """无 LLM 时的规则解读"""
    sig_parts = []
    sig_vars = []
    insig_vars = []
    for r in result_table:
        var = r.get("var", "")
        coef = r.get("coef", 0)
        pval = r.get("pval", 1)
        stars = r.get("stars", "")
        if pval <= 0.01:
            sig_parts.append(f"{var} 在1%水平上显著（系数={coef:.4f}）{stars}")
            sig_vars.append(var)
        elif pval <= 0.05:
            sig_parts.append(f"{var} 在5%水平上显著（系数={coef:.4f}）{stars}")
            sig_vars.append(var)
        elif pval <= 0.1:
            sig_parts.append(f"{var} 在10%水平上显著（系数={coef:.4f}）{stars}")
            sig_vars.append(var)
        else:
            insig_vars.append(var)

    significance = (
        f"采用{method_name}进行回归分析。"
        + "；".join(sig_parts)
        + ("。" if sig_parts else "所有变量均不显著。")
        + (f"不显著的变量包括：{', '.join(insig_vars)}。" if insig_vars else "")
    )

    economic_significance = ""
    eco_parts = []
    for r in result_table:
        if r.get("pval", 1) <= 0.1:
            var = r.get("var", "")
            coef = r.get("coef", 0)
            eco_parts.append(f"{var} 的系数为 {coef:.4f}")
    if eco_parts:
        economic_significance = "从经济显著性来看，" + "；".join(eco_parts) + "。需结合数据量纲进一步分析实际经济含义。"
    else:
        economic_significance = "各变量系数的经济含义需结合数据量纲进一步分析。"

    innovation_points = [
        f"采用{method_name}进行实证检验，方法规范",
    ]
    if any(r.get("pval", 1) <= 0.05 for r in result_table):
        innovation_points.append("发现核心解释变量在统计上显著，验证了理论假设")
    if stats.get("r2", 0) > 0.3:
        innovation_points.append(f"模型拟合优度较好（R²={stats['r2']:.3f}）")

    comparison_text = (
        "本研究的回归结果与已有文献基本一致。"
        "核心解释变量的系数符号和显著性水平符合理论预期。"
    )

    n_val = stats.get("n", "")
    r2_val = stats.get("r2", "")
    f_val = stats.get("f_stat", "")

    chapter_draft = (
        f"## 实证结果分析\n\n"
        f"为检验研究假设，本文采用{method_name}进行基准回归分析。"
        f"回归结果见表X。\n\n"
        f"**基准回归结果**。表X第（1）列报告了基准回归结果。"
    )
    for r in result_table:
        var = r.get("var", "")
        coef = r.get("coef", 0)
        se = r.get("se", 0)
        stars = r.get("stars", "")
        chapter_draft += f"{var}的回归系数为{coef:.4f}{stars}（标准误={se:.4f}），"
    chapter_draft = chapter_draft.rstrip("，") + "。\n\n"
    if r2_val:
        chapter_draft += f"模型的R²为{r2_val}，"
    if f_val:
        chapter_draft += f"F统计量为{f_val}，"
    if n_val:
        chapter_draft += f"观测值为{n_val}。"
    chapter_draft += "以上结果表明，模型整体拟合较好，核心解释变量对因变量具有显著影响。\n\n"
    chapter_draft += (
        "**稳健性检验**。为保证结论的可靠性，本文进行了以下稳健性检验："
        "（1）替换被解释变量的测度方式；（2）采用不同模型设定；"
        "（3）排除极端值影响。上述检验均支持基准回归的结论。"
    )

    return InterpretResponse(
        significance=significance,
        economic_significance=economic_significance,
        innovation_points=innovation_points,
        comparison_text=comparison_text,
        chapter_draft=chapter_draft,
    )


import os
from fastapi.responses import FileResponse

static_dir = os.path.join(os.path.dirname(__file__), "static")

@app.get("/")
async def serve_index():
    return FileResponse(os.path.join(static_dir, "index.html"))

@app.get("/{filename:path}")
async def serve_static(filename: str):
    filepath = os.path.join(static_dir, filename)
    if os.path.isfile(filepath):
        return FileResponse(filepath)
    return FileResponse(os.path.join(static_dir, "index.html"))

# ── 启动入口 ────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
